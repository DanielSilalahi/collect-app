import io
import math
import re
from urllib.parse import quote
from collections import namedtuple
from datetime import date, datetime

import openpyxl
import pytz
from sqlalchemy import MetaData, Table
from fastapi import APIRouter, Depends, Request, UploadFile, File, Form, Query
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, inspect as sqlalchemy_inspect
from core.database import get_db
from core.templates import templates
from models.user import User
from models.customer import Customer
from models.customer_address import CustomerAddress
from models.customer_contact import CustomerContact
from models.customer_import_row import CustomerImportRow
from models.customer_loan import CustomerLoan
from models.activity_log import ActivityLog

router = APIRouter(tags=["Customer Management"])
CustomerImportPayload = namedtuple(
    "CustomerImportPayload",
    ["customer", "loan", "address", "contact", "import_row"],
)


def clean_string(value):
    if value is None:
        return None
    value = str(value).strip()
    if not value or value.lower() == "none":
        return None
    return value


def normalize_text(value):
    cleaned = clean_string(value)
    if not cleaned:
        return None
    lowered = cleaned.lower()
    return re.sub(r"\s+", " ", lowered).strip()


def stringify_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def parse_int(value):
    value = clean_string(value)
    if value is None:
        return None
    try:
        return int(float(value.replace(",", "")))
    except Exception:
        return None


def parse_float(value):
    value = clean_string(value)
    if value is None:
        return None
    try:
        return float(value.replace(",", ""))
    except Exception:
        return None


def parse_decimal(value):
    parsed = parse_float(value)
    if parsed is None:
        return None
    if float(parsed).is_integer():
        return int(parsed)
    return parsed


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = clean_string(value)
    if raw is None:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw[:10], fmt).date()
        except ValueError:
            continue
    return None


FIELD_PARSER_BY_TYPE = {
    "string": clean_string,
    "int": parse_int,
    "float": parse_float,
    "decimal": parse_decimal,
    "date": parse_date,
}

CATEGORY_META = {
    "identity": {
        "label": "Identitas Customer",
        "icon": "bi-person-vcard",
        "description": "Snapshot utama customer untuk identitas, status, dan pencarian dasar.",
    },
    "location": {
        "label": "Alamat & Lokasi",
        "icon": "bi-geo-alt",
        "description": "Alamat utama, detail domisili, dan koordinat lokasi yang terkait customer.",
    },
    "loan": {
        "label": "Pinjaman Aktif",
        "icon": "bi-cash-coin",
        "description": "Field bisnis untuk pinjaman saat ini, outstanding, DPD, pembayaran, dan status collection.",
    },
    "contact": {
        "label": "Kontak Tambahan",
        "icon": "bi-telephone-forward",
        "description": "Kontak darurat atau kontak lain yang berhubungan dengan customer.",
    },
    "import_meta": {
        "label": "Metadata Import",
        "icon": "bi-archive",
        "description": "Metadata partner dan raw field yang disimpan untuk audit hasil upload.",
    },
}

UPLOAD_FIELD_SPECS = [
    ("identity", "full_name", "Nama Lengkap", True, "customer", "full_name", "string"),
    ("identity", "nick_name", "Nama Panggilan", False, "customer", "nick_name", "string"),
    ("identity", "customer_code", "Kode Customer", False, "customer", "customer_code", "string"),
    ("identity", "external_customer_id", "External Customer ID", False, "customer", "external_customer_id", "string"),
    ("identity", "platform_name", "Platform Customer", False, "customer", "platform_name", "string"),
    ("identity", "partner_name", "Nama Partner", False, "customer", "partner_name", "string"),
    ("identity", "nik", "NIK", False, "customer", "nik", "string"),
    ("identity", "birth_date", "Tanggal Lahir", False, "customer", "birth_date", "date"),
    ("identity", "gender", "Gender", False, "customer", "gender", "string"),
    ("identity", "email", "Email Customer", False, "customer", "email", "string"),
    ("identity", "primary_phone", "No. HP Utama", False, "customer", "primary_phone", "string"),
    ("identity", "primary_city", "Kota Snapshot", False, "customer", "primary_city", "string"),
    ("identity", "primary_address_summary", "Ringkasan Alamat Snapshot", False, "customer", "primary_address_summary", "string"),
    ("identity", "status", "Status Customer", False, "customer", "status", "string"),
    ("identity", "sub_status", "Sub Status Customer", False, "customer", "sub_status", "string"),
    ("identity", "current_dpd", "Current DPD Snapshot", False, "customer", "current_dpd", "int"),
    ("identity", "current_total_outstanding", "Current Total Outstanding Snapshot", False, "customer", "current_total_outstanding", "decimal"),
    ("identity", "last_payment_date", "Tanggal Pembayaran Terakhir Snapshot", False, "customer", "last_payment_date", "date"),
    ("identity", "last_payment_amount", "Nominal Pembayaran Terakhir Snapshot", False, "customer", "last_payment_amount", "decimal"),
    ("identity", "customer_notes", "Catatan Customer", False, "customer", "notes", "string"),
    ("location", "address_type", "Tipe Alamat", False, "address", "address_type", "string"),
    ("location", "label", "Label Alamat", False, "address", "label", "string"),
    ("location", "recipient_name", "Nama Penerima Alamat", False, "address", "recipient_name", "string"),
    ("location", "full_address", "Alamat Lengkap", False, "address", "full_address", "string"),
    ("location", "street", "Jalan", False, "address", "street", "string"),
    ("location", "block", "Blok", False, "address", "block", "string"),
    ("location", "house_number", "Nomor Rumah", False, "address", "house_number", "string"),
    ("location", "rt", "RT", False, "address", "rt", "string"),
    ("location", "rw", "RW", False, "address", "rw", "string"),
    ("location", "kelurahan", "Kelurahan", False, "address", "kelurahan", "string"),
    ("location", "kecamatan", "Kecamatan", False, "address", "kecamatan", "string"),
    ("location", "city", "Kota Alamat", False, "address", "city", "string"),
    ("location", "province", "Provinsi", False, "address", "province", "string"),
    ("location", "postal_code", "Kode Pos", False, "address", "postal_code", "string"),
    ("location", "country", "Negara", False, "address", "country", "string"),
    ("location", "landmark", "Patokan", False, "address", "landmark", "string"),
    ("location", "address_note", "Catatan Alamat", False, "address", "address_note", "string"),
    ("location", "residence_status", "Status Tempat Tinggal", False, "address", "residence_status", "string"),
    ("location", "is_primary", "Alamat Utama", False, "address", "is_primary", "int"),
    ("location", "is_active", "Alamat Aktif", False, "address", "is_active", "int"),
    ("location", "lat", "Latitude", False, "address", "lat", "float"),
    ("location", "lng", "Longitude", False, "address", "lng", "float"),
    ("location", "map_url", "Map URL", False, "address", "map_url", "string"),
    ("location", "coordinate_source", "Sumber Koordinat", False, "address", "coordinate_source", "string"),
    ("location", "coordinate_accuracy_meters", "Akurasi Koordinat (meter)", False, "address", "coordinate_accuracy_meters", "float"),
    ("location", "is_location_verified", "Lokasi Terverifikasi", False, "address", "is_location_verified", "int"),
    ("location", "raw_lat", "Raw Latitude", False, "address", "raw_lat", "string"),
    ("location", "raw_lng", "Raw Longitude", False, "address", "raw_lng", "string"),
    ("location", "raw_lat_lng", "Raw Latitude & Longitude", False, "address", "raw_lat_lng", "string"),
    ("location", "raw_map_link", "Raw Map Link", False, "address", "raw_map_link", "string"),
    ("loan", "is_current", "Loan Aktif Saat Ini", False, "loan", "is_current", "int"),
    ("loan", "application_id", "Application ID", False, "loan", "application_id", "string"),
    ("loan", "loan_number", "Loan Number", False, "loan", "loan_number", "string"),
    ("loan", "contract_number", "Contract Number", False, "loan", "contract_number", "string"),
    ("loan", "agreement_number", "Agreement Number", False, "loan", "agreement_number", "string"),
    ("loan", "product_type", "Tipe Produk", False, "loan", "product_type", "string"),
    ("loan", "product_name", "Nama Produk", False, "loan", "product_name", "string"),
    ("loan", "loan_platform_name", "Platform Loan", False, "loan", "platform_name", "string"),
    ("loan", "disbursement_date", "Tanggal Disbursement", False, "loan", "disbursement_date", "date"),
    ("loan", "first_due_date", "Tanggal Jatuh Tempo Pertama", False, "loan", "first_due_date", "date"),
    ("loan", "due_date", "Tanggal Jatuh Tempo", False, "loan", "due_date", "date"),
    ("loan", "last_due_date", "Tanggal Jatuh Tempo Terakhir", False, "loan", "last_due_date", "date"),
    ("loan", "maturity_date", "Tanggal Maturity", False, "loan", "maturity_date", "date"),
    ("loan", "tenor", "Tenor", False, "loan", "tenor", "int"),
    ("loan", "installment_number", "Installment Number", False, "loan", "installment_number", "int"),
    ("loan", "remaining_installment_count", "Sisa Cicilan", False, "loan", "remaining_installment_count", "int"),
    ("loan", "payment_frequency", "Frekuensi Pembayaran", False, "loan", "payment_frequency", "string"),
    ("loan", "loan_amount", "Loan Amount", False, "loan", "loan_amount", "decimal"),
    ("loan", "principal_amount", "Principal Amount", False, "loan", "principal_amount", "decimal"),
    ("loan", "interest_amount", "Interest Amount", False, "loan", "interest_amount", "decimal"),
    ("loan", "admin_fee_amount", "Admin Fee Amount", False, "loan", "admin_fee_amount", "decimal"),
    ("loan", "penalty_amount", "Penalty Amount", False, "loan", "penalty_amount", "decimal"),
    ("loan", "insurance_fee_amount", "Insurance Fee Amount", False, "loan", "insurance_fee_amount", "decimal"),
    ("loan", "other_fee_amount", "Other Fee Amount", False, "loan", "other_fee_amount", "decimal"),
    ("loan", "installment_amount", "Installment Amount", False, "loan", "installment_amount", "decimal"),
    ("loan", "outstanding_principal", "Outstanding Principal", False, "loan", "outstanding_principal", "decimal"),
    ("loan", "outstanding_interest", "Outstanding Interest", False, "loan", "outstanding_interest", "decimal"),
    ("loan", "outstanding_penalty", "Outstanding Penalty", False, "loan", "outstanding_penalty", "decimal"),
    ("loan", "outstanding_fee", "Outstanding Fee", False, "loan", "outstanding_fee", "decimal"),
    ("loan", "total_outstanding", "Total Outstanding", False, "loan", "total_outstanding", "decimal"),
    ("loan", "remaining_balance", "Remaining Balance", False, "loan", "remaining_balance", "decimal"),
    ("loan", "overdue_days", "Overdue Days", False, "loan", "overdue_days", "int"),
    ("loan", "days_past_due", "Days Past Due", False, "loan", "days_past_due", "int"),
    ("loan", "dpd_bucket", "DPD Bucket", False, "loan", "dpd_bucket", "string"),
    ("loan", "aging_bucket", "Aging Bucket", False, "loan", "aging_bucket", "string"),
    ("loan", "bucket_code", "Bucket Code", False, "loan", "bucket_code", "string"),
    ("loan", "loan_status", "Loan Status", False, "loan", "loan_status", "string"),
    ("loan", "billing_status", "Billing Status", False, "loan", "billing_status", "string"),
    ("loan", "collection_stage", "Collection Stage", False, "loan", "collection_stage", "string"),
    ("loan", "risk_segment", "Risk Segment", False, "loan", "risk_segment", "string"),
    ("loan", "risk_score", "Risk Score", False, "loan", "risk_score", "string"),
    ("loan", "loan_last_payment_date", "Tanggal Pembayaran Terakhir Loan", False, "loan", "last_payment_date", "date"),
    ("loan", "loan_last_payment_amount", "Nominal Pembayaran Terakhir Loan", False, "loan", "last_payment_amount", "decimal"),
    ("loan", "last_payment_channel", "Channel Pembayaran Terakhir", False, "loan", "last_payment_channel", "string"),
    ("loan", "last_payment_reference", "Referensi Pembayaran Terakhir", False, "loan", "last_payment_reference", "string"),
    ("loan", "paid_amount_total", "Total Paid Amount", False, "loan", "paid_amount_total", "decimal"),
    ("loan", "payment_status", "Payment Status", False, "loan", "payment_status", "string"),
    ("loan", "promise_to_pay_date", "Tanggal Promise To Pay", False, "loan", "promise_to_pay_date", "date"),
    ("loan", "promise_to_pay_amount", "Nominal Promise To Pay", False, "loan", "promise_to_pay_amount", "decimal"),
    ("loan", "promise_to_pay_status", "Status Promise To Pay", False, "loan", "promise_to_pay_status", "string"),
    ("loan", "broken_ptp_count", "Broken PTP Count", False, "loan", "broken_ptp_count", "int"),
    ("loan", "settlement_offer_amount", "Settlement Offer Amount", False, "loan", "settlement_offer_amount", "decimal"),
    ("loan", "settlement_expiry_date", "Settlement Expiry Date", False, "loan", "settlement_expiry_date", "date"),
    ("loan", "minimum_payment_amount", "Minimum Payment Amount", False, "loan", "minimum_payment_amount", "decimal"),
    ("contact", "contact_type", "Tipe Kontak", False, "contact", "contact_type", "string"),
    ("contact", "contact_role", "Peran Kontak", False, "contact", "contact_role", "string"),
    ("contact", "contact_name", "Nama Kontak", False, "contact", "name", "string"),
    ("contact", "relationship_label", "Relationship", False, "contact", "relationship_label", "string"),
    ("contact", "contact_phone_number", "No. HP Kontak", False, "contact", "phone_number", "string"),
    ("contact", "contact_email", "Email Kontak", False, "contact", "email", "string"),
    ("contact", "contact_is_primary", "Kontak Utama", False, "contact", "is_primary", "int"),
    ("contact", "priority_order", "Priority Order", False, "contact", "priority_order", "int"),
    ("contact", "is_whatsapp", "Nomor WhatsApp", False, "contact", "is_whatsapp", "int"),
    ("contact", "contact_is_active", "Kontak Aktif", False, "contact", "is_active", "int"),
    ("contact", "is_verified", "Kontak Terverifikasi", False, "contact", "is_verified", "int"),
    ("contact", "is_valid", "Kontak Valid", False, "contact", "is_valid", "int"),
    ("contact", "verification_source", "Sumber Verifikasi", False, "contact", "verification_source", "string"),
    ("contact", "contact_notes", "Catatan Kontak", False, "contact", "notes", "string"),
    ("import_meta", "source_partner_name", "Source Partner Name", False, "import_row", "source_partner_name", "string"),
    ("import_meta", "source_partner_code", "Source Partner Code", False, "import_row", "source_partner_code", "string"),
    ("import_meta", "source_file_name", "Source File Name", False, "import_row", "source_file_name", "string"),
    ("import_meta", "source_sheet_name", "Source Sheet Name", False, "import_row", "source_sheet_name", "string"),
    ("import_meta", "source_row_number", "Source Row Number", False, "import_row", "source_row_number", "int"),
    ("import_meta", "mapping_profile_name", "Mapping Profile Name", False, "import_row", "mapping_profile_name", "string"),
    ("import_meta", "import_version", "Import Version", False, "import_row", "import_version", "string"),
    ("import_meta", "import_status", "Import Status", False, "import_row", "import_status", "string"),
    ("import_meta", "import_error_flag", "Import Error Flag", False, "import_row", "import_error_flag", "int"),
    ("import_meta", "import_error_message", "Import Error Message", False, "import_row", "import_error_message", "string"),
    ("import_meta", "raw_customer_name", "Raw Customer Name", False, "import_row", "raw_customer_name", "string"),
    ("import_meta", "raw_nik", "Raw NIK", False, "import_row", "raw_nik", "string"),
    ("import_meta", "raw_phone", "Raw Phone", False, "import_row", "raw_phone", "string"),
    ("import_meta", "raw_phone_2", "Raw Phone 2", False, "import_row", "raw_phone_2", "string"),
    ("import_meta", "raw_address", "Raw Address", False, "import_row", "raw_address", "string"),
    ("import_meta", "raw_city", "Raw City", False, "import_row", "raw_city", "string"),
    ("import_meta", "raw_due_date", "Raw Due Date", False, "import_row", "raw_due_date", "string"),
    ("import_meta", "raw_disbursement_date", "Raw Disbursement Date", False, "import_row", "raw_disbursement_date", "string"),
    ("import_meta", "raw_loan_amount", "Raw Loan Amount", False, "import_row", "raw_loan_amount", "string"),
    ("import_meta", "raw_installment_amount", "Raw Installment Amount", False, "import_row", "raw_installment_amount", "string"),
    ("import_meta", "raw_outstanding_amount", "Raw Outstanding Amount", False, "import_row", "raw_outstanding_amount", "string"),
    ("import_meta", "raw_overdue_days", "Raw Overdue Days", False, "import_row", "raw_overdue_days", "string"),
    ("import_meta", "raw_platform_name", "Raw Platform Name", False, "import_row", "raw_platform_name", "string"),
    ("import_meta", "raw_status", "Raw Status", False, "import_row", "raw_status", "string"),
]

UPLOAD_FIELD_DEFINITIONS = [
    {
        "category": category,
        "key": key,
        "label": label,
        "required": required,
        "target": target,
        "attr": attr,
        "type": field_type,
    }
    for category, key, label, required, target, attr, field_type in UPLOAD_FIELD_SPECS
]

FIELD_DEFINITION_BY_KEY = {field["key"]: field for field in UPLOAD_FIELD_DEFINITIONS}
LEGACY_FIELD_ALIASES = {
    "emergency_contact_name": "contact_name",
    "emergency_contact_phone": "contact_phone_number",
}
TARGET_MODEL_BY_KEY = {
    "customer": Customer,
    "address": CustomerAddress,
    "loan": CustomerLoan,
    "contact": CustomerContact,
    "import_row": CustomerImportRow,
}
TARGET_ATTR_COLUMN_BY_KEY = {
    target: {
        attr.key: attr.columns[0].name
        for attr in model.__mapper__.column_attrs
        if attr.columns
    }
    for target, model in TARGET_MODEL_BY_KEY.items()
}


def get_val(row, mapping, field_name):
    source_key = mapping.get(field_name)
    if source_key is None:
        return None
    if isinstance(row, dict):
        return row.get(source_key)
    return None


def parse_mapped_value(row, mapping, field_name):
    field_definition = FIELD_DEFINITION_BY_KEY.get(field_name)
    if not field_definition:
        return None
    parser = FIELD_PARSER_BY_TYPE[field_definition["type"]]
    return parser(get_val(row, mapping, field_name))


def get_runtime_columns_by_target(db: Session):
    inspector = sqlalchemy_inspect(db.get_bind())
    return {
        target: {column["name"] for column in inspector.get_columns(model.__table__.name)}
        for target, model in TARGET_MODEL_BY_KEY.items()
    }


def get_runtime_tables(db: Session):
    metadata = MetaData()
    bind = db.get_bind()
    return {
        target: Table(model.__table__.name, metadata, autoload_with=bind)
        for target, model in TARGET_MODEL_BY_KEY.items()
    }


def filter_field_definitions_by_runtime_columns(runtime_columns_by_target):
    visible_fields = []
    for field in UPLOAD_FIELD_DEFINITIONS:
        column_name = TARGET_ATTR_COLUMN_BY_KEY[field["target"]].get(field["attr"])
        if column_name and column_name in runtime_columns_by_target.get(field["target"], set()):
            visible_fields.append(field)
    return visible_fields


def serialize_model_for_runtime_insert(model_instance, target, runtime_columns_by_target):
    values = {}
    runtime_columns = runtime_columns_by_target.get(target, set())
    for attr, column_name in TARGET_ATTR_COLUMN_BY_KEY[target].items():
        if column_name not in runtime_columns:
            continue
        if column_name == "id":
            continue
        values[column_name] = getattr(model_instance, attr)

    if target == "customer":
        if "name" in runtime_columns and not values.get("name"):
            values["name"] = (
                getattr(model_instance, "full_name", None)
                or getattr(model_instance, "nick_name", None)
                or "-"
            )
        if "address" in runtime_columns and not values.get("address"):
            values["address"] = getattr(model_instance, "primary_address_summary", None)
        if "phone" in runtime_columns and not values.get("phone"):
            values["phone"] = getattr(model_instance, "primary_phone", None)
        if "outstanding_amount" in runtime_columns and not values.get("outstanding_amount"):
            values["outstanding_amount"] = getattr(model_instance, "current_total_outstanding", None)
        if "overdue_days" in runtime_columns and not values.get("overdue_days"):
            values["overdue_days"] = getattr(model_instance, "current_dpd", None)
        if "is_deleted" in runtime_columns and values.get("is_deleted") is None:
            values["is_deleted"] = 0
    return values


def build_category_groups(field_definitions=None):
    field_definitions = field_definitions or UPLOAD_FIELD_DEFINITIONS
    groups = []
    for category_key, meta in CATEGORY_META.items():
        groups.append(
            {
                "key": category_key,
                "label": meta["label"],
                "icon": meta["icon"],
                "description": meta["description"],
                "fields": [
                    field for field in field_definitions if field["category"] == category_key
                ],
            }
        )
    return groups


def resolve_mapping_value(row, mapping, field_name, *aliases):
    for candidate in (field_name, *aliases):
        if candidate not in mapping:
            continue
        if candidate in FIELD_DEFINITION_BY_KEY:
            return parse_mapped_value(row, mapping, candidate)
        return clean_string(get_val(row, mapping, candidate))
    return None


def build_customer_import_payload(row, mapping, batch_code, field_definitions=None, runtime_columns_by_target=None):
    field_definitions = field_definitions or UPLOAD_FIELD_DEFINITIONS
    normalized_mapping = {LEGACY_FIELD_ALIASES.get(key, key): value for key, value in mapping.items()}
    values_by_target = {key: {} for key in ("customer", "address", "loan", "contact", "import_row")}
    raw_values_by_key = {}

    def supports_attr(target, attr):
        if runtime_columns_by_target is None:
            return True
        column_name = TARGET_ATTR_COLUMN_BY_KEY[target].get(attr)
        return column_name in runtime_columns_by_target.get(target, set())

    def set_value(target, attr, value):
        if not supports_attr(target, attr):
            return
        values_by_target[target][attr] = value

    for field in field_definitions:
        if field["key"] not in normalized_mapping:
            continue
        raw_values_by_key[field["key"]] = get_val(row, normalized_mapping, field["key"])
        set_value(
            field["target"],
            field["attr"],
            parse_mapped_value(
                row,
                normalized_mapping,
                field["key"],
            ),
        )

    full_name = values_by_target["customer"].get("full_name")
    nik = values_by_target["customer"].get("nik")
    full_address = values_by_target["address"].get("full_address")
    address_city = values_by_target["address"].get("city")
    customer_platform_name = values_by_target["customer"].get("platform_name")
    loan_platform_name = values_by_target["loan"].get("platform_name")

    if not values_by_target["customer"].get("primary_address_summary"):
        set_value("customer", "primary_address_summary", full_address)
    if not values_by_target["customer"].get("primary_city"):
        set_value("customer", "primary_city", address_city)
    if not values_by_target["customer"].get("status"):
        set_value("customer", "status", "new")
    set_value("customer", "upload_batch", batch_code)
    set_value("customer", "search_name", normalize_text(full_name))
    set_value("customer", "search_nik", normalize_text(nik))
    if values_by_target["customer"].get("current_dpd") is None:
        set_value("customer", "current_dpd", values_by_target["loan"].get("overdue_days"))
    if values_by_target["customer"].get("current_total_outstanding") is None:
        set_value("customer", "current_total_outstanding", values_by_target["loan"].get("total_outstanding"))
    if values_by_target["customer"].get("last_payment_date") is None:
        set_value("customer", "last_payment_date", values_by_target["loan"].get("last_payment_date"))
    if values_by_target["customer"].get("last_payment_amount") is None:
        set_value("customer", "last_payment_amount", values_by_target["loan"].get("last_payment_amount"))

    if values_by_target["loan"].get("is_current") is None:
        set_value("loan", "is_current", 1)
    if not values_by_target["loan"].get("platform_name"):
        set_value("loan", "platform_name", customer_platform_name)

    if not values_by_target["address"].get("address_type"):
        set_value("address", "address_type", "home")
    if not values_by_target["address"].get("full_address"):
        set_value(
            "address",
            "full_address",
            full_address or values_by_target["customer"].get("primary_address_summary") or "-",
        )
    if values_by_target["address"].get("is_primary") is None:
        set_value("address", "is_primary", 1)
    if values_by_target["address"].get("is_active") is None:
        set_value("address", "is_active", 1)

    contact_name = resolve_mapping_value(row, normalized_mapping, "contact_name", "emergency_contact_name")
    contact_phone = resolve_mapping_value(
        row,
        normalized_mapping,
        "contact_phone_number",
        "emergency_contact_phone",
    )
    if contact_name is not None:
        set_value("contact", "name", contact_name)
    if contact_phone is not None:
        set_value("contact", "phone_number", contact_phone)
    if not values_by_target["contact"].get("contact_type"):
        set_value("contact", "contact_type", "phone")
    if not values_by_target["contact"].get("contact_role"):
        set_value("contact", "contact_role", "emergency")
    if values_by_target["contact"].get("is_primary") is None:
        set_value("contact", "is_primary", 0)
    if values_by_target["contact"].get("is_active") is None:
        set_value("contact", "is_active", 1)
    if values_by_target["contact"].get("is_valid") is None:
        set_value("contact", "is_valid", 1)
    if values_by_target["contact"].get("is_whatsapp") is None:
        set_value("contact", "is_whatsapp", 0)
    if values_by_target["contact"].get("is_verified") is None:
        set_value("contact", "is_verified", 0)

    set_value("import_row", "upload_batch", batch_code)
    if not values_by_target["import_row"].get("import_status"):
        set_value("import_row", "import_status", "imported")
    if values_by_target["import_row"].get("import_error_flag") is None:
        set_value("import_row", "import_error_flag", 0)
    set_value("import_row", "raw_payload", dict(row))
    if not values_by_target["import_row"].get("raw_customer_name"):
        set_value("import_row", "raw_customer_name", stringify_value(raw_values_by_key.get("full_name")) or full_name)
    if not values_by_target["import_row"].get("raw_nik"):
        set_value("import_row", "raw_nik", stringify_value(raw_values_by_key.get("nik")) or nik)
    if not values_by_target["import_row"].get("raw_phone"):
        set_value(
            "import_row",
            "raw_phone",
            stringify_value(raw_values_by_key.get("primary_phone")) or values_by_target["customer"].get("primary_phone"),
        )
    if not values_by_target["import_row"].get("raw_address"):
        set_value("import_row", "raw_address", stringify_value(raw_values_by_key.get("full_address")) or full_address)
    if not values_by_target["import_row"].get("raw_city"):
        set_value(
            "import_row",
            "raw_city",
            stringify_value(raw_values_by_key.get("city")) or address_city or values_by_target["customer"].get("primary_city"),
        )
    if not values_by_target["import_row"].get("raw_due_date"):
        set_value("import_row", "raw_due_date", stringify_value(raw_values_by_key.get("due_date")))
    if not values_by_target["import_row"].get("raw_disbursement_date"):
        set_value("import_row", "raw_disbursement_date", stringify_value(raw_values_by_key.get("disbursement_date")))
    if not values_by_target["import_row"].get("raw_loan_amount"):
        set_value("import_row", "raw_loan_amount", stringify_value(raw_values_by_key.get("loan_amount")))
    if not values_by_target["import_row"].get("raw_installment_amount"):
        set_value("import_row", "raw_installment_amount", stringify_value(raw_values_by_key.get("installment_amount")))
    if not values_by_target["import_row"].get("raw_outstanding_amount"):
        set_value("import_row", "raw_outstanding_amount", stringify_value(raw_values_by_key.get("total_outstanding")))
    if not values_by_target["import_row"].get("raw_overdue_days"):
        set_value("import_row", "raw_overdue_days", stringify_value(raw_values_by_key.get("overdue_days")))
    if not values_by_target["import_row"].get("raw_lat"):
        set_value(
            "import_row",
            "raw_lat",
            stringify_value(raw_values_by_key.get("lat")) or stringify_value(values_by_target["address"].get("lat")),
        )
    if not values_by_target["import_row"].get("raw_lng"):
        set_value(
            "import_row",
            "raw_lng",
            stringify_value(raw_values_by_key.get("lng")) or stringify_value(values_by_target["address"].get("lng")),
        )
    if not values_by_target["import_row"].get("raw_lat_lng"):
        set_value(
            "import_row",
            "raw_lat_lng",
            stringify_value(raw_values_by_key.get("raw_lat_lng")) or values_by_target["address"].get("raw_lat_lng"),
        )
    if not values_by_target["import_row"].get("raw_platform_name"):
        set_value(
            "import_row",
            "raw_platform_name",
            stringify_value(raw_values_by_key.get("platform_name")) or customer_platform_name or loan_platform_name,
        )
    if not values_by_target["import_row"].get("raw_status"):
        set_value("import_row", "raw_status", stringify_value(raw_values_by_key.get("status")) or values_by_target["customer"].get("status"))

    customer = Customer(**values_by_target["customer"])
    loan = CustomerLoan(**values_by_target["loan"])
    address = CustomerAddress(**values_by_target["address"])

    contact = None
    if any(
        values_by_target["contact"].get(field_name) is not None
        for field_name in ("name", "phone_number", "email", "relationship_label")
    ):
        contact = CustomerContact(**values_by_target["contact"])

    import_row = CustomerImportRow(**values_by_target["import_row"])

    return CustomerImportPayload(
        customer=customer,
        loan=loan,
        address=address,
        contact=contact,
        import_row=import_row,
    )


def _require_admin(request: Request, db: Session):
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    user = db.query(User).filter(User.id == user_id, User.role == "admin").first()
    return user


@router.get("/customers", response_class=HTMLResponse, name="customers")
def customer_list_batches(
    request: Request,
    db: Session = Depends(get_db),
):
    """Customer batch list overview."""
    current_user = _require_admin(request, db)
    if not current_user:
        return RedirectResponse("/login", status_code=302)

    from sqlalchemy import case, func

    batches_query = db.query(
        Customer.upload_batch,
        func.count(Customer.id).label("total_customers"),
        func.sum(case((Customer.assigned_agent_id != None, 1), else_=0)).label("assigned_count"),
        func.max(Customer.created_at).label("upload_date")
    ).filter(Customer.is_deleted == 0).group_by(Customer.upload_batch).order_by(func.max(Customer.created_at).desc()).all()

    agents = db.query(User).filter(User.role == "agent", User.is_active == True).all()

    # Get success/error from query params
    success = request.query_params.get("success")
    error = request.query_params.get("error")

    return templates.TemplateResponse(
        request,
        "customers/batches.html",
        {
            "current_user": current_user,
            "batches": batches_query,
            "agents": agents,
            "success": success,
            "error": error,
        },
    )

@router.get("/customers/batch/{batch_code}", response_class=HTMLResponse, name="customer_batch_detail")
def customer_batch_detail(
    request: Request,
    batch_code: str,
    db: Session = Depends(get_db),
    page: int = Query(default=1, ge=1),
    status: str = Query(default=None),
    agent_id: str = Query(default=None),
    search: str = Query(default=None),
):
    """Customer list page for a specific batch."""
    current_user = _require_admin(request, db)
    if not current_user:
        return RedirectResponse("/login", status_code=302)

    per_page = 20
    from sqlalchemy.orm import joinedload, selectinload
    from models.collection import Collection
    query = db.query(Customer).options(
        joinedload(Customer.agent),
        joinedload(Customer.current_loan),
        selectinload(Customer.addresses),
        selectinload(Customer.contacts),
        joinedload(Customer.collections).joinedload(Collection.agent)
    ).filter(Customer.is_deleted == 0)
    
    if batch_code.lower() == "manual":
        query = query.filter(Customer.upload_batch == None)
    else:
        query = query.filter(Customer.upload_batch == batch_code)

    if status:
        query = query.filter(Customer.status == status)
    if agent_id and agent_id.isdigit():
        query = query.filter(Customer.assigned_agent_id == int(agent_id))
    if search:
        query = query.filter(Customer.full_name.ilike(f"%{search}%"))

    total = query.count()
    total_pages = math.ceil(total / per_page) if total > 0 else 1
    if page > total_pages:
        page = total_pages

    customers = (
        query.order_by(Customer.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    agents = db.query(User).filter(User.role == "agent", User.is_active == True).all()

    success = request.query_params.get("success")
    error = request.query_params.get("error")

    return templates.TemplateResponse(
        request,
        "customers/list.html",
        {
            "current_user": current_user,
            "customers": customers,
            "agents": agents,
            "page": page,
            "total_pages": total_pages,
            "total": total,
            "status_filter": status,
            "agent_filter": agent_id,
            "search": search or "",
            "success": success,
            "error": error,
            "batch_code": batch_code
        },
    )


@router.post("/customers/upload")
async def upload_customers(
    request: Request,
    file: UploadFile = File(...),
    agent_id: int = Form(None),
    db: Session = Depends(get_db),
):
    """Step 1: Upload customers from Excel file and show mapping UI."""
    current_user = _require_admin(request, db)
    if not current_user:
        return RedirectResponse("/login", status_code=302)

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return RedirectResponse(
            f"/customers?error={quote('Format file harus .xlsx')}",
            status_code=302,
        )

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        
        # Read the first row to get headers
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell).strip() if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
            break
        
        wb.close()

        if not headers:
            return RedirectResponse(
                f"/customers?error={quote('File kosong atau tidak memiliki header')}",
                status_code=302,
            )

        # Save file temporarily
        import os, uuid
        temp_dir = "static/uploads/temp"
        os.makedirs(temp_dir, exist_ok=True)
        temp_filename = f"{uuid.uuid4().hex}.xlsx"
        temp_filepath = os.path.join(temp_dir, temp_filename)
        
        with open(temp_filepath, "wb") as f:
            f.write(contents)

        runtime_columns_by_target = get_runtime_columns_by_target(db)
        expected_fields = filter_field_definitions_by_runtime_columns(runtime_columns_by_target)
        category_groups = build_category_groups(expected_fields)

        return templates.TemplateResponse(
            request,
            "customers/upload_mapping.html",
            {
                "current_user": current_user,
                "headers": headers,
                "expected_fields": expected_fields,
                "category_groups": category_groups,
                "temp_filename": temp_filename,
                "agent_id": agent_id or "",
            },
        )

    except Exception as e:
        return RedirectResponse(
            f"/customers?error={quote(f'Gagal memproses file: {str(e)}')}",
            status_code=302,
        )

@router.post("/customers/upload/process")
async def process_customers_upload(
    request: Request,
    temp_filename: str = Form(...),
    agent_id: str = Form(None),
    db: Session = Depends(get_db),
):
    """Step 2: Process mapping and insert data."""
    current_user = _require_admin(request, db)
    if not current_user:
        return RedirectResponse("/login", status_code=302)

    import os
    temp_filepath = os.path.join("static/uploads/temp", temp_filename)
    if not os.path.exists(temp_filepath):
        return RedirectResponse(
            f"/customers?error={quote('File temporary tidak ditemukan. Silahkan upload ulang.')}",
            status_code=302,
        )

    try:
        form_data = await request.form()
        
        # Get mapping configuration
        # e.g., form_data["mapped_name"] = "0" (index string of the column)
        mapping = {}
        for key in form_data.keys():
            if key.startswith("mapped_"):
                field_name = key.replace("mapped_", "")
                col_idx_str = form_data.get(key)
                if col_idx_str and col_idx_str.isdigit():
                    mapping[field_name] = int(col_idx_str)

        if "full_name" not in mapping:
             return RedirectResponse(
                f"/customers?error={quote('Mapping kolom Nama wajib diisi.')}",
                status_code=302,
             )

        # Parse the actual file
        wb = openpyxl.load_workbook(temp_filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return RedirectResponse(
                f"/customers?error={quote('File kosong')}",
                status_code=302,
            )

        jakarta = pytz.timezone("Asia/Jakarta")
        batch_code = f"UPLOAD_{datetime.now(jakarta).strftime('%Y%m%d_%H%M%S')}"
        count = 0

        agent_id_val = int(agent_id) if agent_id and agent_id.isdigit() else None
        runtime_columns_by_target = get_runtime_columns_by_target(db)
        runtime_tables_by_target = get_runtime_tables(db)
        visible_field_definitions = filter_field_definitions_by_runtime_columns(runtime_columns_by_target)

        headers = [str(cell).strip() if cell is not None else f"Column_{i}" for i, cell in enumerate(rows[0])]

        for row in rows[1:]:
            if all(v is None for v in row):
                continue

            row_payload = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            mapped_headers = {
                field_name: headers[col_idx]
                for field_name, col_idx in mapping.items()
                if col_idx < len(headers)
            }

            payload = build_customer_import_payload(
                row=row_payload,
                mapping=mapped_headers,
                batch_code=batch_code,
                field_definitions=visible_field_definitions,
                runtime_columns_by_target=runtime_columns_by_target,
            )

            if not payload.customer.full_name:
                continue

            payload.customer.assigned_agent_id = agent_id_val
            customer_insert = serialize_model_for_runtime_insert(
                payload.customer,
                "customer",
                runtime_columns_by_target,
            )
            customer_result = db.execute(
                runtime_tables_by_target["customer"].insert().values(**customer_insert)
            )
            customer_id = customer_result.inserted_primary_key[0]

            payload.loan.customer_id = customer_id
            db.add(payload.loan)
            db.flush()

            customer_update = {}
            if "current_loan_id" in runtime_columns_by_target["customer"]:
                customer_update["current_loan_id"] = payload.loan.id
            if "current_dpd" in runtime_columns_by_target["customer"]:
                customer_update["current_dpd"] = payload.loan.overdue_days
            if "current_total_outstanding" in runtime_columns_by_target["customer"]:
                customer_update["current_total_outstanding"] = payload.loan.total_outstanding
            if "loan_number" in runtime_columns_by_target["customer"]:
                customer_update["loan_number"] = payload.loan.loan_number
            if "platform_name" in runtime_columns_by_target["customer"] and not customer_insert.get("platform_name"):
                customer_update["platform_name"] = payload.loan.platform_name or payload.customer.platform_name
            if "outstanding_amount" in runtime_columns_by_target["customer"]:
                customer_update["outstanding_amount"] = payload.loan.total_outstanding
            if "due_date" in runtime_columns_by_target["customer"]:
                customer_update["due_date"] = payload.loan.due_date
            if "overdue_days" in runtime_columns_by_target["customer"]:
                customer_update["overdue_days"] = payload.loan.overdue_days

            payload.address.customer_id = customer_id
            db.add(payload.address)

            if payload.contact:
                payload.contact.customer_id = customer_id
                db.add(payload.contact)
                if "emergency_contact_1_name" in runtime_columns_by_target["customer"] and not customer_update.get("emergency_contact_1_name"):
                    customer_update["emergency_contact_1_name"] = payload.contact.name
                if "emergency_contact_1_phone" in runtime_columns_by_target["customer"] and not customer_update.get("emergency_contact_1_phone"):
                    customer_update["emergency_contact_1_phone"] = payload.contact.phone_number

            payload.import_row.customer_id = customer_id
            db.add(payload.import_row)
            if customer_update:
                db.execute(
                    runtime_tables_by_target["customer"]
                    .update()
                    .where(runtime_tables_by_target["customer"].c.id == customer_id)
                    .values(**customer_update)
                )
            count += 1

        # Cleanup
        try:
            os.remove(temp_filepath)
        except:
            pass

        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action="upload_customer",
            detail=f"Upload dynamically {count} customers (batch: {batch_code})",
        )
        db.add(log)
        db.commit()

        return RedirectResponse(
            f"/customers?success={quote(f'Berhasil upload {count} customer (batch: {batch_code})')}",
            status_code=302,
        )

    except Exception as e:
        db.rollback()
        return RedirectResponse(
            f"/customers?error={quote(f'Gagal proses data: {str(e)}')}",
            status_code=302,
        )


@router.post("/customers/assign")
def assign_customer(
    request: Request,
    customer_ids: str = Form(...),
    agent_id: int = Form(...),
    db: Session = Depends(get_db),
):
    """Assign customers to an agent."""
    current_user = _require_admin(request, db)
    if not current_user:
        return RedirectResponse("/login", status_code=302)

    ids = [int(x.strip()) for x in customer_ids.split(",") if x.strip().isdigit()]
    if not ids:
        return RedirectResponse(
            f"/customers?error={quote('Tidak ada customer yang dipilih')}",
            status_code=302,
        )

    agent = db.query(User).filter(User.id == agent_id, User.role == "agent").first()
    if not agent:
        return RedirectResponse(
            f"/customers?error={quote('Agent tidak ditemukan')}",
            status_code=302,
        )

    updated = (
        db.query(Customer)
        .filter(Customer.id.in_(ids))
        .update({"assigned_agent_id": agent_id}, synchronize_session="fetch")
    )

    log = ActivityLog(
        user_id=current_user.id,
        action="assign_customer",
        detail=f"Assign {updated} customers ke agent {agent.name}",
    )
    db.add(log)
    db.commit()

    referer = request.headers.get("referer") or "/customers"
    if "?" in referer:
        referer = referer.split("?")[0]

    return RedirectResponse(
        f"{referer}?success={quote(f'Berhasil assign {updated} customer ke {agent.name}')}",
        status_code=302,
    )


@router.post("/customers/bulk-delete")
def bulk_delete_customers(
    request: Request,
    customer_ids: str = Form(...),
    db: Session = Depends(get_db),
):
    """Bulk delete customers."""
    current_user = _require_admin(request, db)
    if not current_user:
        return RedirectResponse("/login", status_code=302)

    ids = [int(x.strip()) for x in customer_ids.split(",") if x.strip().isdigit()]
    if not ids:
        return RedirectResponse(
            f"/customers?error={quote('Tidak ada customer yang dipilih untuk dihapus')}",
            status_code=302,
        )

    updated = (
        db.query(Customer)
        .filter(Customer.id.in_(ids))
        .update({"is_deleted": 1}, synchronize_session="fetch")
    )

    log = ActivityLog(
        user_id=current_user.id,
        action="delete_bulk_customer",
        detail=f"Menghapus secara lembut (soft delete) {updated} customers",
    )
    db.add(log)
    db.commit()

    referer = request.headers.get("referer") or "/customers"
    if "?" in referer:
        referer = referer.split("?")[0]

    return RedirectResponse(
        f"{referer}?success={quote(f'Berhasil menghapus {updated} data customer')}",
        status_code=302,
    )


@router.post("/customers/batch/{batch_code}/delete")
def delete_batch(
    request: Request,
    batch_code: str,
    db: Session = Depends(get_db),
):
    """Delete all customers in a batch."""
    current_user = _require_admin(request, db)
    if not current_user:
        return RedirectResponse("/login", status_code=302)

    updated = (
        db.query(Customer)
        .filter(Customer.upload_batch == batch_code)
        .update({"is_deleted": 1}, synchronize_session="fetch")
    )

    log = ActivityLog(
        user_id=current_user.id,
        action="delete_batch",
        detail=f"Hapus batch (soft delete) {batch_code} ({updated} customers)",
    )
    db.add(log)
    db.commit()

    return RedirectResponse(
        f"/customers?success={quote(f'Batch {batch_code} dan {updated} data di dalamnya berhasil dihapus')}",
        status_code=302,
    )


@router.post("/customers/{customer_id}/delete")
def delete_customer(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Delete a customer."""
    current_user = _require_admin(request, db)
    if not current_user:
        return RedirectResponse("/login", status_code=302)

    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return RedirectResponse(
            f"/customers?error={quote('Customer tidak ditemukan')}",
            status_code=302,
        )

    name = customer.full_name
    customer.is_deleted = 1

    log = ActivityLog(
        user_id=current_user.id,
        action="delete_customer",
        detail=f"Hapus customer (soft delete): {name}",
    )
    db.add(log)
    db.commit()

    referer = request.headers.get("referer") or "/customers"
    if "?" in referer:
        referer = referer.split("?")[0]

    return RedirectResponse(
        f"{referer}?success={quote(f'Customer {name} berhasil dihapus')}",
        status_code=302,
    )
